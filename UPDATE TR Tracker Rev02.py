import re
import io
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
import pdfplumber
import openpyxl
import streamlit as st


# ─────────────────────────────────────────────────────────────
# LOOKUP TABLES
# ─────────────────────────────────────────────────────────────

FULL_NAMES = {
    "WA 115.1: Particle Size Distribution":
        "WA 115.1: Particle Size Distribution: Sieving and Decantation Method",
    "WA 141.1: Determination of the California":
        "WA 141.1: Determination of the California Bearing Ratio of a Soil: Standard Laboratory Method for a Remoulded Specimen",
    "WA 115.2: Particle Size Distribution: Abbreviated":
        "WA 115.2: Particle Size Distribution: Abbreviated Method for Coarse and Medium Grained Soils",
    "WA 133.1: Dry Density/Moisture Content":
        "WA 133.1: Dry Density/Moisture Content Relationship: Modified Compaction Fine and Medium Grained Soils",
    "WA 324.2: Determination of Field Density":
        "WA 324.2: Determination of Field Density: Nuclear Method",
    "Construction Moisture Content (WA 110.1)":
        "Construction Moisture Content (WA 110.1) - Convection Oven Method",
    "Construction Moisture Content (WA 110.2)":
        "Construction Moisture Content (WA 110.2) - Microwave Oven Method",
}

MATCH_ALIASES = {
    "wa testing field density 6 ndm sites x 6 mdd":
        "field density package 6 ndm sites x 6 mdd",
    "wa testing field density 6 ndm sites x 3 mdd":
        "field density package 6 ndm sites x 3 mdd",
    "wa testing field density 3 ndm sites x 3 mdd":
        "field density package 3 ndm sites x 3 mdd",
    "wa testing field density 3 ndm sites x 2 mdd":
        "field density package 3 ndm sites x 2 mdd",
    "wa testing field density 6 ndm sites x 2 mdd":
        "field density package 6 ndm sites x 2 mdd",
    "wa testing field density 9 ndm sites x 3 mdd":
        "field density package 9 ndm sites x 3 mdd",
    "wa 110 136 1 dryback single layer mrwa":
        "wa 136 1 moisture ratio dryback",
    "wa 123 1 linear shrink mrwa":
        "wa 123 1 linear shrinkage",
    "wa 122 1 plasticity index mrwa":
        "wa 122 1 plasticity index",
    "wa 115 1 particle size distribution mrwa":
        "wa 115 1 particle size distribution sieving and decantation method",
    "wa 141 1 scbr":
        "wa 141 1 determination of the california bearing ratio of a soil standard laboratory method for a remoulded specimen",
    "wa 144 1 foreign material mrwa":
        "wa 144 1 foreign material",
    "as 1289 6 3 3 perth sand penetrometer":
        "as 1289 6 3 3 soil strength and consolidation tests",
    "as 1289 4 1 1 organic content soil chemical test":
        "as 1289 4 1 1 organic testing",
    "wa 133 1 modified compaction mrwa":
        "wa 133 1 dry density moisture content relationship modified compaction fine and medium grained soils",
    "wa 115 2 abbreviated psd":
        "wa 115 2 particle size distribution abbreviated method for coarse and medium grained soils",
    "as 1289 6 3 2 dynamic cone penetrometer":
        "as 1289 6 3 2 soil strength and consolidation tests determination of the penetration resistance of a soil dynamic cone penetrometer test",
    "wa 120 2 atterberg 4 point cone penetrometer":
        "wa 120 2 liquid limit plastic limit and linear shrinkage index point method",
}

FIELD_DENSITY_MAP = {
    (2, 6): "Field Density Package - 6 NDM Sites x 2 MDD",
    (2, 3): "Field Density Package - 3 NDM Sites x 2 MDD",
    (3, 3): "Field Density Package - 3 NDM Sites x 3 MDD",
    (3, 6): "Field Density Package - 6 NDM Sites x 3 MDD",
    (3, 9): "Field Density Package - 9 NDM Sites x 3 MDD",
    (6, 6): "Field Density Package - 6 NDM Sites x 6 MDD",
}


# ─────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────

def normalize_method_name(name):
    """
    Expands abbreviated test method names to their full standardised titles.

    Iterates over the FULL_NAMES lookup table and checks whether any known
    abbreviated key is a substring of the provided name string. If a match is
    found, the full name is returned in its place. If no match is found, the
    original name is returned unchanged.

    Args:
        name (str): Raw test method name extracted from the PDF text.

    Returns:
        str: The full standardised test method name, or the original if no
             match is found.
    """
    for key, full in FULL_NAMES.items(): #.items() gives tuples of the dict key/name pairs
        if key in name:
            return full
    return name


def normalize_match_text(text):
    """
    Converts a test method name or claimed test string into a normalised,
    lowercase, punctuation-free form suitable for fuzzy comparison.

    Processing steps:
      1. Returns an empty string for NaN/None values.
      2. Converts to lowercase and strips leading/trailing whitespace.
      3. Removes punctuation characters: . : - ( ) / ,
      4. Collapses consecutive whitespace into a single space.
      5. Strips again after collapsing.
      6. Applies the MATCH_ALIASES lookup to remap known variant phrasings
         to their canonical form.

    Args:
        text (str | float): Raw test name from either the PDF or claim sheet.

    Returns:
        str: Normalised text ready for word-overlap comparison.
    """
    if pd.isna(text): #step 1
        return ""

    text = str(text).lower().strip() #step 2
    text = re.sub(r"[\.\:\-\(\)\/\,]", " ", text) #step 3
    text = re.sub(r"\s+", " ", text) #step 4
    text = text.strip() #step 5

    if text in MATCH_ALIASES:
        text = MATCH_ALIASES[text] #step 6

    return text


def words_match(a, b):
    """
    Determines whether two normalised test method strings refer to the same
    test using a multi-strategy fuzzy matching approach.

    Matching strategies (applied in order, returns True on first match):
      1. Exact equality of the two strings.
      2. One string is a full substring of the other (handles cases where one
         description is a truncated version of the other).
      3. Word-set overlap ratio: computes the number of shared words divided
         by the size of the larger word set. Returns True if this ratio meets
         or exceeds the 0.6 threshold.

    Args:
        a (str): First normalised test method string.
        b (str): Second normalised test method string.

    Returns:
        bool: True if the two strings are considered a match, False otherwise.
    """
    a_words = set(a.split())
    b_words = set(b.split())

    if a == b:
        return True

    if a in b or b in a:
        return True

    overlap = len(a_words & b_words)
    total = max(len(a_words), len(b_words))

    if total == 0:
        return False

    return (overlap / total) >= 0.6


def normalize_tr(tr_value):
    """
    Extracts and normalises a TR (test request) number from various raw string
    formats into a clean integer-equivalent string.

    Processing steps:
      1. Returns None for NaN/None values.
      2. Strips whitespace from the string representation.
      3. Uses a regex to extract the first contiguous digit sequence found.
      4. Converts the matched digits to int (removing leading zeros) and back
         to string.
      5. Returns None if no digit sequence is found.

    Args:
        tr_value (str | int | float): Raw TR value as read from a PDF or
                                      Excel cell.

    Returns:
        str | None: Normalised TR number as a string (e.g. "12345"), or None
                    if extraction fails.
    """
    if pd.isna(tr_value):
        return None

    tr_value = str(tr_value).strip()
    match = re.search(r"(\d+)", tr_value)

    if not match:
        return None

    return str(int(match.group(1)))


def replace_field_density(methods):
    """
    Consolidates individual WA 133.1 (MDD) and WA 134.1 / WA 324.2 (NDM)
    test method entries into a single named Field Density Package string.

    The input, methods, is a list looking like this:
        methods = [
        ("WA 133.1 MDD", 2),
        ("WA 134.1 NDM", 8),
        ("WA 133.1 Modified", 1)
        ]
    So when I call m,c in methods I am referencing method and count.
    
    Logic:
      1. Counts the total number of WA 133.1 tests across all method entries
         (these represent MDD compaction tests).
      2. Counts the combined total of WA 134.1 and WA 324.2 tests (these
         represent NDM field density sites).
      3. Looks up the (MDD count, NDM count) key in FIELD_DENSITY_MAP to find
         the corresponding package name.
      4. If a matching package is found, removes all WA 133.1, WA 134.1, and
         WA 324.2 entries from the list and appends a single entry for the
         package with a count of 1.
      5. If no match is found in the map, the original list is returned
         unchanged.

    Args:
        methods (list[tuple[str, int]]): List of (method_name, count) tuples
                                         extracted from a single TR page.

    Returns:
        list[tuple[str, int]]: Updated list with field density entries
                               consolidated, or the original list if no
                               consolidation applies.
    """
    count_133 = sum(c for m, c in methods if "WA 133.1" in m)
    count_134 = sum(c for m, c in methods if "WA 134.1" in m)
    count_324 = sum(c for m, c in methods if "WA 324.2" in m)
    count_134_combined = count_134 + count_324

    key = (count_133, count_134_combined)

    if key in FIELD_DENSITY_MAP:
        methods = [
            (m, c)
            for m, c in methods
            if "WA 133.1" not in m
            and "WA 134.1" not in m
            and "WA 324.2" not in m
        ]
        methods.append((FIELD_DENSITY_MAP[key], 1))

    return methods


def process_pdf(pdf_file):

    """
    Parses a single CivilPro test request PDF and extracts structured test
    data from every page that contains a TR number.

    For each page:
      1. Extracts raw text; skips pages with no text or no TR number.
      2. Normalises the TR number via normalize_tr().
      3. Extracts Lot No. and derives Lot Type from its first two characters.
      4. Extracts the 'When Req'd' date using a regex on the date line.
      5. Reads the 'Location Method' field to determine parsing strategy:
         - 'tester locates': matches lines of the form
           "<count> WA|AS <code>: <description>" and accumulates methods.
         - 'location specified': matches numbered lines of the form
           "<n>-<n> WA|AS <code>: <description>", counts occurrences per
           method using a Counter, and strips trailing price artefacts.
      6. Passes the collected methods list through replace_field_density() to
         consolidate field density packages.
      7. Appends one row per method (or a blank method row if none found).

    Args:
        pdf_file: A file-like object (BytesIO) of the PDF to process.

    Returns:
        pd.DataFrame: DataFrame with columns:
            TR, When Req'd (date), Test method, No. tests, Lot no., Lot Type
    """
    rows = []

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()

            if not text:
                continue

            tr_match = re.search(r"TR:\s*(\d+)", text)

            if not tr_match:
                continue

            tr = normalize_tr(tr_match.group(1))

            lot_match = re.search(r"Lot No:\s*([A-Za-z0-9\-]+)", text)
            lot_no = lot_match.group(1) if lot_match else ""
            lot_type = lot_no[:2] if lot_no else ""

            date_match = re.search(
                r"When Req'd\s+\w+,\s+(\d{2}\s+\w+\s+\d{4})",
                text
            )
            when_reqd = (
                pd.to_datetime(date_match.group(1), format="%d %b %Y").date()
                if date_match
                else None
            )

            location_method = None
            loc_match = re.search(r"Location Method:\s*([A-Za-z ]+)", text)
            if loc_match:
                location_method = loc_match.group(1).strip().lower()

            methods = []

            if location_method == "tester locates":
                for count, std, code, desc in re.findall(
                    r"(\d+)\s+(WA|AS)\s+([\d\.]+):\s+([^\n]+)",
                    text
                ):
                    method_name = f"{std} {code}: {desc.strip()}"
                    method_name = normalize_method_name(method_name)
                    methods.append((method_name, int(count)))

            elif location_method == "location specified":
                numbered = re.findall(
                    r"\d+-\d+\s+(WA|AS)\s+([\d\.]+):\s+([^\n]+)",
                    text
                )

                if numbered:
                    counter = Counter()
                    for std, code, desc in numbered:
                        clean_desc = re.sub(r"\s+0\.0+.*$", "", desc).strip()
                        method_name = f"{std} {code}: {clean_desc}"
                        method_name = normalize_method_name(method_name)
                        counter[method_name] += 1
                    for method, cnt in counter.items():
                        methods.append((method, cnt))

            methods = replace_field_density(methods)

            if not methods:
                rows.append([tr, when_reqd, "", None, lot_no, lot_type])
            else:
                for method, count in methods:
                    rows.append([tr, when_reqd, method, count, lot_no, lot_type])

    df = pd.DataFrame(
        rows,
        columns=[
            "TR",
            "When Req'd (date)",
            "Test method",
            "No. tests",
            "Lot no.",
            "Lot Type",
        ],
    )
    df["No. tests"] = pd.to_numeric(df["No. tests"], errors="coerce")
    return df


def parse_claim_sheet(file):
    """
    Reads the 'Detail' tab from a CivilPro claim Excel workbook and extracts
    a flat table of claimed test items with their associated TR and date.

    Parsing behaviour:
      - Starts reading from row 2 (row 1 is usually the header. Watch for errors due to hardcode.).
      - Skips rows where all relevant cells (ticket, TR, code, method, qty,
        date) are None.
      - Tracks the most recently seen TR and date values across rows, since
        the Excel layout uses merged/sparse cells where TR and date only
        appear on the first row of a group.
      - Skips rows where both code and method are None (non-test rows).
      - Normalises TR via normalize_tr().
      - Concatenates code and method into a single 'Claimed test' string,
        handling cases where one of the two may be absent.
      - Raises ValueError if the workbook does not contain a 'Detail' sheet.

    Args:
        file: A file-like object (BytesIO) of the Excel workbook to parse.

    Returns:
        pd.DataFrame: DataFrame with columns:
            TR, Date, Claimed test, Qty
    """
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)

    if "Detail" not in wb.sheetnames:
        raise ValueError("Claim sheet missing Detail tab")

    ws = wb["Detail"]
    rows = []
    current_tr = None
    current_date = None

    for row in ws.iter_rows(min_row=2):
        ticket = row[1].value
        tr_val = row[2].value
        code = row[3].value
        method = row[4].value
        qty = row[5].value
        date_val = row[6].value

        if all(v is None for v in [ticket, tr_val, code, method, qty, date_val]):
            continue

        if tr_val is not None:
            current_tr = str(tr_val).strip()

        if date_val is not None:
            if isinstance(date_val, datetime):
                current_date = date_val.date()
            else:
                current_date = date_val

        if code is None and method is None:
            continue

        tr_num = normalize_tr(current_tr)

        rows.append({
            "TR": tr_num,
            "Date": current_date,
            "Claimed test": (
                f"{code} - {method}".strip(" -")
                if code and method
                else (method or code or "")
            ),
            "Qty": qty,
        })

    return pd.DataFrame(rows)


def build_pivot(cp_df):
    """
    Produces a summary pivot of total test counts grouped by test method.

    Groups the CivilPro DataFrame by the 'Test method' column (including NaN
    values so blank method rows are not silently dropped), sums the
    'No. tests' column for each group, and sorts the result in descending
    order of total tests so the most common methods appear at the top.

    Args:
        cp_df (pd.DataFrame): The combined CivilPro DataFrame produced by
                              processing all uploaded PDFs.

    Returns:
        pd.DataFrame: DataFrame with columns:
            Test method, No. tests
        Sorted descending by No. tests.
    """
    return (
        cp_df.groupby("Test method", dropna=False)["No. tests"]
        .sum()
        .reset_index()
        .sort_values("No. tests", ascending=False)
    )


def build_discrepancies(cp_df, claim_df):
    """
    Compares CivilPro test request data against claim sheet data on a
    per-TR basis and produces a row-level discrepancy report.

    Matching logic per TR:
      - If a TR exists only in CivilPro (not in the claim sheet): all CP
        methods are flagged as "NOT IN CLAIM".
      - If a TR exists only in the claim sheet (not in CivilPro): all claimed
        items are flagged as "NO CIVILPRO MATCH".
      - If a TR exists in both: each claimed item is normalised via
        normalize_match_text() and compared against each CivilPro method
        using words_match(). The first CP method that passes words_match()
        is used:
          - If quantities match: flagged as "Yes".
          - If quantities differ: flagged as "QUANTITY MISMATCH".
          - If no CP method matches: flagged as "TEST MISMATCH", and all CP
            methods for that TR are concatenated with "; " as context.

    TR ordering: all TRs from both sources are sorted numerically (non-digit
    TRs sort last via the key lambda).

    Args:
        cp_df (pd.DataFrame): CivilPro DataFrame from processing PDFs.
        claim_df (pd.DataFrame): Claim sheet DataFrame from processing Excels.

    Returns:
        pd.DataFrame: DataFrame with columns:
            TR, Date, Match, Claimed test, Inv. Quantity,
            CivilPro test, CivilPro Quantity
    """
    discrepancy_rows = []

    claim_by_tr = (
        claim_df.dropna(subset=["TR"])
        .groupby("TR", group_keys=False)
        .apply(lambda g: g[["Date", "Claimed test", "Qty"]].to_dict("records"))
        .to_dict()
    )

    cp_by_tr = (
        cp_df.dropna(subset=["TR"])
        .groupby("TR")
        .apply(lambda g: g[["Test method", "No. tests"]].to_dict("records"))
        .to_dict()
    )

    all_trs = sorted(
        set(list(claim_by_tr.keys()) + list(cp_by_tr.keys())),
        key=lambda x: int(x) if x.isdigit() else 0,
    )

    for tr in all_trs:
        claimed_items = claim_by_tr.get(tr, [])
        cp_items = cp_by_tr.get(tr, [])

        if not claimed_items:
            for cp in cp_items:
                discrepancy_rows.append({
                    "TR": tr,
                    "Date": None,
                    "Match": "NOT IN CLAIM",
                    "Claimed test": "",
                    "Inv. Quantity": "",
                    "CivilPro test": cp["Test method"],
                    "CivilPro Quantity": cp["No. tests"],
                })

        elif not cp_items:
            for item in claimed_items:
                discrepancy_rows.append({
                    "TR": tr,
                    "Date": item.get("Date"),
                    "Match": "NO CIVILPRO MATCH",
                    "Claimed test": item.get("Claimed test"),
                    "Inv. Quantity": item.get("Qty"),
                    "CivilPro test": "",
                    "CivilPro Quantity": "",
                })

        else:
            for item in claimed_items:
                claimed_test = item.get("Claimed test", "")
                inv_qty = item.get("Qty")
                date = item.get("Date")

                normalized_claim = normalize_match_text(claimed_test)
                matched_cp = None

                for cp in cp_items:
                    normalized_cp = normalize_match_text(cp["Test method"])
                    if words_match(normalized_claim, normalized_cp):
                        matched_cp = cp
                        break

                if matched_cp:
                    cp_qty = matched_cp["No. tests"]
                    qty_match = (
                        pd.to_numeric(inv_qty, errors="coerce")
                        == pd.to_numeric(cp_qty, errors="coerce")
                    )
                    discrepancy_rows.append({
                        "TR": tr,
                        "Date": date,
                        "Match": "Yes" if qty_match else "QUANTITY MISMATCH",
                        "Claimed test": claimed_test,
                        "Inv. Quantity": inv_qty,
                        "CivilPro test": matched_cp["Test method"],
                        "CivilPro Quantity": cp_qty,
                    })
                else:
                    discrepancy_rows.append({
                        "TR": tr,
                        "Date": date,
                        "Match": "TEST MISMATCH",
                        "Claimed test": claimed_test,
                        "Inv. Quantity": inv_qty,
                        "CivilPro test": "; ".join(
                            cp["Test method"] for cp in cp_items
                        ),
                        "CivilPro Quantity": "; ".join(
                            str(cp["No. tests"]) for cp in cp_items
                        ),
                    })

    disc_df = pd.DataFrame(discrepancy_rows)

    return disc_df[[
        "TR", "Date", "Match", "Claimed test",
        "Inv. Quantity", "CivilPro test", "CivilPro Quantity",
    ]]


def build_excel_output(cp_df, pivot_df, disc_df):
    """
    Serialises the three result DataFrames into a single multi-sheet Excel
    workbook held in memory and returns the raw bytes.

    Sheet layout:
      - 'All tests': Full row-level CivilPro test data (cp_df).
      - 'Tests summary': Pivot table of total tests by method (pivot_df).
      - 'Discrepancies': Row-level comparison report (disc_df).

    Post-processing applied to every sheet:
      - Auto-fits column widths by iterating over all cells, measuring the
        string length of each value, and setting column width to
        min(max_length + 2, 60) to avoid excessively wide columns.

    Args:
        cp_df (pd.DataFrame): CivilPro all-tests DataFrame.
        pivot_df (pd.DataFrame): Test method summary pivot DataFrame.
        disc_df (pd.DataFrame): Discrepancy report DataFrame.

    Returns:
        bytes: Raw bytes of the completed .xlsx workbook, ready to be
               written to disk or served as a Streamlit download.
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        cp_df.to_excel(writer, sheet_name="All tests", index=False)
        pivot_df.to_excel(writer, sheet_name="Tests summary", index=False)
        disc_df.to_excel(writer, sheet_name="Discrepancies", index=False)

        for sheet_name in writer.sheets:
            ws_sheet = writer.sheets[sheet_name]
            for col_cells in ws_sheet.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws_sheet.column_dimensions[col_letter].width = min(
                    max_length + 2, 60
                )

    return buffer.getvalue()


def check_duplicate_trs(claim_df):
    """
    Scans the claim DataFrame for TR numbers that appear more than once and
    returns a summary Series for display as a warning.

    A TR is considered a duplicate if its normalised TR value appears on
    more than one row in the claim DataFrame. This can indicate that two
    different claim Excel files contain entries for the same TR, or that a
    single file has duplicate rows.

    Args:
        claim_df (pd.DataFrame): Combined claim DataFrame from all uploaded
                                 claim sheets.

    Returns:
        pd.Series: Value counts for TR numbers that appear more than once,
                   indexed by TR. Empty Series if no duplicates exist.
    """
    counts = claim_df["TR"].value_counts()
    return counts[counts > 1]


# ─────────────────────────────────────────────────────────────
# STREAMLIT UI AND MAIN EXECUTION
# ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="TR Claim Reconciliation",
    page_icon="📋",
    layout="wide",
)

st.title("Western Geo TR Claim Reconciliation")
st.markdown('''WHAT IS THIS?  \n
    This tool does two things:  \n
    1. Scrapes one or more CivilPro Test Request PDFs into a combined Excel table.  \n
    2. Compares those TRs against a WGLS claim sheet and shows any discrepancies.  \n

    HOW TO USE:
    1. Upload one or more Test Request PDFs (from CivilPro → Printer → "Test Requests (pdf)").
    2. Click "Process CivilPro PDFs" to combine them all into one TR list.
    3. Upload your WGLS Claim Sheet (.xlsx) to the second box.
    4. The app will match TRs and show a discrepancy report.
    5. Download the final Excel file (All tests / Tests summary / Discrepancies tabs).

    If you have any questions or it breaks, please ask Kieran.''')


col1, col2 = st.columns(2)

with col1:
    st.subheader("CivilPro PDFs")
    pdf_uploads = st.file_uploader(
        "Upload one or more CivilPro TR PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        key="pdf_uploader",
    )

with col2:
    st.subheader("Claim Excel Sheets")
    claim_uploads = st.file_uploader(
        "Upload one or more claim Excel files",
        type=["xlsx"],
        accept_multiple_files=True,
        key="claim_uploader",
    )

run_button = st.button(
    "▶ Run Reconciliation",
    type="primary",
    disabled=(not pdf_uploads or not claim_uploads),
)

if run_button:

    # ── Process PDFs ──────────────────────────────────────────
    st.divider()
    st.subheader("Processing PDFs…")

    all_pdf_frames = []
    pdf_progress = st.progress(0, text="Reading PDFs…")

    for i, pdf_file in enumerate(pdf_uploads):
        df_part = process_pdf(io.BytesIO(pdf_file.read()))
        all_pdf_frames.append(df_part)
        pdf_progress.progress(
            (i + 1) / len(pdf_uploads),
            text=f"Read {i + 1} / {len(pdf_uploads)} PDFs"
        )

    cp_df = pd.concat(all_pdf_frames, ignore_index=True)
    pdf_progress.empty()
    st.success(
        f"Processed **{len(pdf_uploads)}** PDF(s) — "
        f"**{len(cp_df)}** total test rows extracted."
    )

    # ── Process Claim Sheets ──────────────────────────────────
    st.subheader("Processing Claim Sheets…")

    all_claim_frames = []
    claim_errors = []
    claim_progress = st.progress(0, text="Reading claim sheets…")

    for i, claim_file in enumerate(claim_uploads):
        try:
            df_claim = parse_claim_sheet(io.BytesIO(claim_file.read()))
            df_claim["Source File"] = claim_file.name
            all_claim_frames.append(df_claim)
        except Exception as e:
            claim_errors.append(f"{claim_file.name}: {e}")

        claim_progress.progress(
            (i + 1) / len(claim_uploads),
            text=f"Read {i + 1} / {len(claim_uploads)} claim sheets"
        )

    claim_progress.empty()

    if claim_errors:
        for err in claim_errors:
            st.error(f"❌ {err}")

    if not all_claim_frames:
        st.error("No valid claim sheets could be parsed. Aborting.")
        st.stop()

    claim_df = pd.concat(all_claim_frames, ignore_index=True)

    st.success(
        f"Processed **{len(claim_uploads)}** claim sheet(s) — "
        f"**{len(claim_df)}** total claim rows extracted."
    )

    # ── Duplicate TR warning ──────────────────────────────────
    duplicates = check_duplicate_trs(claim_df)

    if len(duplicates) > 0:
        with st.expander("Duplicate TRs found in claim sheets", expanded=True):
            for tr, count in duplicates.items():
                st.warning(f"TR {tr} appears {count} times")

    # ── Build outputs ─────────────────────────────────────────
    st.subheader("Building Report…")

    with st.spinner("Generating pivot and discrepancy tables…"):
        pivot_df = build_pivot(cp_df)
        disc_df = build_discrepancies(cp_df, claim_df)
        excel_bytes = build_excel_output(cp_df, pivot_df, disc_df)

    st.success("Report ready!")

    # ── Summary metrics ───────────────────────────────────────
    st.divider()
    st.subheader("Summary")

    total_trs = cp_df["TR"].nunique()
    matched = (disc_df["Match"] == "Yes").sum()
    mismatches = disc_df[disc_df["Match"] != "Yes"]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Unique TRs (PDFs)", total_trs)
    m2.metric("Matched rows", matched)
    m3.metric("Discrepancy rows", len(mismatches))
    m4.metric("Total claim rows", len(claim_df))

    # ── Preview tabs ──────────────────────────────────────────
    tab1, tab2, tab3 = st.tabs([
        "Discrepancies",
        "Tests Summary",
        "All Tests (PDF extract)",
    ])

    with tab1:
        st.dataframe(disc_df, use_container_width=True, height=500)

    with tab2:
        st.dataframe(pivot_df, use_container_width=True)

    with tab3:
        st.dataframe(cp_df, use_container_width=True, height=500)

    # ── Download ──────────────────────────────────────────────
    st.divider()
    st.download_button(
        label="Download Excel Report",
        data=excel_bytes,
        file_name="tr_claim_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
