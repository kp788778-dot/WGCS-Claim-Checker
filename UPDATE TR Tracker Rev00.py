import streamlit as st
import pdfplumber
import pandas as pd
import re
from collections import Counter
from io import BytesIO
import openpyxl
from datetime import datetime

st.title("TR Claim Checker")
st.write('''WHAT IS THIS?
This tool does two things:
1. Scrapes one or more CivilPro Test Request PDFs into a combined Excel table.
2. Compares those TRs against a WGLS claim sheet and shows any discrepancies.

HOW TO USE:
1. Upload one or more Test Request PDFs (from CivilPro → Printer → "Test Requests (pdf)").
2. Click "Process CivilPro PDFs" to combine them all into one TR list.
3. Upload your WGLS Claim Sheet (.xlsx) to the second box.
4. The app will match TRs and show a discrepancy report.
5. Download the final Excel file (All tests / Tests summary / Discrepancies tabs).

If you have any questions or it breaks, please ask Kieran.''')

# ─────────────────────────────────────────────────────────────
# Full-name normalisation
# ─────────────────────────────────────────────────────────────
full_names = {
    "WA 115.1: Particle Size Distribution": "WA 115.1: Particle Size Distribution: Sieving and Decantation Method",
    "WA 141.1: Determination of the California": "WA 141.1: Determination of the California Bearing Ratio of a Soil: Standard Laboratory Method for a Remoulded Specimen",
    "WA 115.2: Particle Size Distribution: Abbreviated": "WA 115.2: Particle Size Distribution: Abbreviated Method for Coarse and Medium Grained Soils",
    "WA 133.1: Dry Density/Moisture Content": "WA 133.1: Dry Density/Moisture Content Relationship: Modified Compaction Fine and Medium Grained Soils",
    "WA 324.2: Determination of Field Density": "WA 324.2: Determination of Field Density: Nuclear Method",
    "Construction Moisture Content (WA 110.1)": "Construction Moisture Content (WA 110.1) - Convection Oven Method",
    "Construction Moisture Content (WA 110.2)": "Construction Moisture Content (WA 110.2) - Microwave Oven Method",
}


def normalize_method_name(name):
    for key, full in full_names.items():
        if key in name:
            return full
    return name


def replace_field_density(methods):
    count_133 = sum(c for m, c in methods if "WA 133.1" in m)
    count_134 = sum(c for m, c in methods if "WA 134.1" in m)
    count_324 = sum(c for m, c in methods if "WA 324.2" in m)
    count_134_combined = count_134 + count_324

    mapping = {
        (2, 6): "WA Testing - Field Density-6 NDM Sites x 2 MDD",
        (2, 3): "WA Testing - Field Density-3 NDM Sites x 2 MDD",
        (3, 3): "WA Testing - Field Density-3 NDM Sites x 3 MDD",
        (3, 6): "WA Testing - Field Density-6 NDM Sites x 3 MDD",
        (3, 9): "WA Testing - Field Density-9 NDM Sites x 3 MDD",
        (6, 6): "WA Testing - Field Density-6 NDM Sites x 6 MDD",
    }

    key = (count_133, count_134_combined)
    if key in mapping:
        methods = [
            (m, c)
            for m, c in methods
            if "WA 133.1" not in m and "WA 134.1" not in m and "WA 324.2" not in m
        ]
        methods.append((mapping[key], 1))
    return methods


def process_pdf(uploaded_file):
    rows = []
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            tr_match = re.search(r"TR:\s*(\d+)", text)
            if not tr_match:
                continue
            tr = tr_match.group(1)

            lot_match = re.search(r"Lot No:\s*([A-Za-z0-9\-]+)", text)
            lot_no = lot_match.group(1) if lot_match else ""
            lot_type = lot_no[:2] if lot_no else ""

            date_match = re.search(r"When Req'd\s+\w+,\s+(\d{2}\s+\w+\s+\d{4})", text)
            when_reqd = (
                pd.to_datetime(date_match.group(1), format="%d %b %Y").date()
                if date_match else None
            )

            location_method = None
            loc_match = re.search(r"Location Method:\s*([A-Za-z ]+)", text)
            if loc_match:
                location_method = loc_match.group(1).strip().lower()

            methods = []
            if location_method == "tester locates":
                for count, std, code, desc in re.findall(
                    r"(\d+)\s+(WA|AS)\s+([\d\.]+):\s+([^\n]+)", text
                ):
                    method_name = f"{std} {code}: {desc.strip()}"
                    method_name = normalize_method_name(method_name)
                    methods.append((method_name, int(count)))

            elif location_method == "location specified":
                numbered = re.findall(
                    r"\d+-\d+\s+(WA|AS)\s+([\d\.]+):\s+([^\n]+)", text
                )
                if numbered:
                    counter = Counter()
                    for std, code, desc in numbered:
                        clean_desc = re.sub(r"\s+0\.0+.*$", "", desc).strip()
                        method_name = f"{std} {code}: {clean_desc}"
                        method_name = normalize_method_name(method_name)
                        counter[method_name] += 1
                    for method, cnt in counter.items():
                        methods.append((method, cnt))

            methods = replace_field_density(methods)

            if not methods:
                rows.append([tr, when_reqd, "", None, lot_no, lot_type])
            else:
                for method, count in methods:
                    rows.append([tr, when_reqd, method, count, lot_no, lot_type])

    return pd.DataFrame(
        rows,
        columns=["TR", "When Req'd (date)", "Test method", "No. tests", "Lot no.", "Lot Type"],
    )


# ─────────────────────────────────────────────────────────────
# Parse claim sheet (Detail tab)
# ─────────────────────────────────────────────────────────────
def parse_claim_sheet(uploaded_file):
    """
    Parse the Detail tab.
    Columns: B=Ticket#, C=Test Request (TR###), D=Code, E=Test Method, F=Qty, G=Date
    TR and Date only appear on the first line of each group – carry them forward.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    if "Detail" not in wb.sheetnames:
        st.error("Claim sheet does not have a 'Detail' tab. Please check your file.")
        return None
    ws = wb["Detail"]

    rows = []
    current_tr = None
    current_date = None

    for row in ws.iter_rows(min_row=2):  # skip header
        ticket   = row[1].value  # col B
        tr_val   = row[2].value  # col C
        code     = row[3].value  # col D
        method   = row[4].value  # col E
        qty      = row[5].value  # col F
        date_val = row[6].value  # col G

        # Skip fully empty rows
        if all(v is None for v in [ticket, tr_val, code, method, qty, date_val]):
            continue

        # Update carry-forward values when a new TR block starts
        if tr_val is not None:
            current_tr = str(tr_val).strip()
        if date_val is not None:
            if isinstance(date_val, datetime):
                current_date = date_val.date()
            else:
                current_date = date_val

        if code is None and method is None:
            continue

        # Extract numeric TR number
        tr_num = None
        if current_tr:
            m = re.search(r"(\d+)", current_tr)
            if m:
                tr_num = m.group(1)

        rows.append({
            "TR": tr_num,
            "Date": current_date,
            "Claimed test": f"{code} - {method}".strip(" -") if code and method else (method or code or ""),
            "Qty": qty,
        })

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────
# UI – Section 1: CivilPro PDFs
# ─────────────────────────────────────────────────────────────
st.markdown("---")
st.header("Step 1 – CivilPro Test Request PDFs")

uploaded_pdfs = st.file_uploader(
    "Upload one or more Test Request PDFs",
    type="pdf",
    accept_multiple_files=True,
    key="pdf_uploader",
)

if "civilpro_df" not in st.session_state:
    st.session_state["civilpro_df"] = None

if st.button("Process CivilPro PDFs") and uploaded_pdfs:
    all_frames = []
    with st.spinner("Processing PDFs…"):
        for pdf_file in uploaded_pdfs:
            df_part = process_pdf(pdf_file)
            all_frames.append(df_part)
    if all_frames:
        combined = pd.concat(all_frames, ignore_index=True)
        st.session_state["civilpro_df"] = combined
        st.success(f"Processed {len(uploaded_pdfs)} PDF(s) — {len(combined)} rows total.")
        st.dataframe(combined)
    else:
        st.warning("No data could be extracted from the uploaded PDFs.")

elif st.session_state["civilpro_df"] is not None:
    st.info("Using previously processed CivilPro data. Upload new PDFs and click the button to reprocess.")
    st.dataframe(st.session_state["civilpro_df"])

# ─────────────────────────────────────────────────────────────
# UI – Section 2: Claim Sheet
# ─────────────────────────────────────────────────────────────
st.markdown("---")
st.header("Step 2 – WGLS Claim Sheet")

claim_file = st.file_uploader(
    "Upload WGLS Claim Sheet (.xlsx)",
    type=["xlsx"],
    key="claim_uploader",
)

claim_df = None
if claim_file:
    with st.spinner("Reading claim sheet…"):
        claim_df = parse_claim_sheet(claim_file)
    if claim_df is not None:
        st.success(f"Claim sheet loaded — {len(claim_df)} line items.")
        st.dataframe(claim_df.head(30))

# ─────────────────────────────────────────────────────────────
# UI – Section 3: Date filter + Download
# ─────────────────────────────────────────────────────────────
st.markdown("---")
st.header("Step 3 – Download Excel Report")

cp_df = st.session_state.get("civilpro_df")

if cp_df is not None:
    st.subheader("Tests Summary Date Filter")
    st.write("Filter the date range for the *Tests summary* tab (e.g. 26 Mar – 26 Apr for a monthly claim).")

    cp_df["When Req'd (date)"] = pd.to_datetime(cp_df["When Req'd (date)"], errors="coerce")
    valid_dates = cp_df["When Req'd (date)"].dropna()

    if len(valid_dates) > 0:
        min_date = valid_dates.min().date()
        max_date = valid_dates.max().date()
    else:
        min_date = max_date = None

    col1, col2 = st.columns(2)
    with col1:
        filter_start = st.date_input("Summary start date", value=min_date, key="start_date")
    with col2:
        filter_end = st.date_input("Summary end date", value=max_date, key="end_date")

    if st.button("Generate & Download Excel"):
        cp_df["No. tests"] = pd.to_numeric(cp_df["No. tests"], errors="coerce")

        # Filtered summary
        mask = pd.Series([True] * len(cp_df))
        if filter_start:
            mask &= cp_df["When Req'd (date)"].dt.date >= filter_start
        if filter_end:
            mask &= cp_df["When Req'd (date)"].dt.date <= filter_end

        filtered_df = cp_df[mask]
        pivot_df = (
            filtered_df.groupby("Test method", dropna=False)["No. tests"]
            .sum()
            .reset_index()
            .sort_values("No. tests", ascending=False)
        )

        # Discrepancies
        discrepancy_rows = []
        if claim_df is not None and "TR" in claim_df.columns:
            # Group claim by TR → list of claimed tests
            claim_by_tr = (
                claim_df.dropna(subset=["TR"])
                .groupby("TR")
                .apply(lambda g: g[["Date", "Claimed test"]].to_dict("records"))
                .to_dict()
            )
            cp_by_tr = (
                cp_df.dropna(subset=["TR"])
                .groupby("TR")["Test method"]
                .apply(list)
                .to_dict()
            )

            all_trs = sorted(
                set(list(claim_by_tr.keys()) + list(cp_by_tr.keys())),
                key=lambda x: int(x) if x.isdigit() else 0,
            )

            for tr in all_trs:
                claimed_items = claim_by_tr.get(tr, [])
                cp_methods = cp_by_tr.get(tr, [])

                if not claimed_items:
                    # In CivilPro but not in claim
                    for method in cp_methods:
                        discrepancy_rows.append({
                            "TR": tr,
                            "Date": None,
                            "Match": "NO MATCH",
                            "Claimed test": "NOT IN CLAIM",
                            "CivilPro test": method,
                        })
                elif not cp_methods:
                    # In claim but not in CivilPro
                    for item in claimed_items:
                        discrepancy_rows.append({
                            "TR": tr,
                            "Date": item.get("Date"),
                            "Match": "NO MATCH",
                            "Claimed test": item.get("Claimed test"),
                            "CivilPro test": "NO MATCH",
                        })
                else:
                    # Both exist – simple presence check
                    # Use first date from claim
                    first_date = claimed_items[0].get("Date") if claimed_items else None
                    claimed_tests = [item.get("Claimed test", "") for item in claimed_items]
                    # Build a simple keyword matcher: check if claimed test code appears in any CP method
                    for item in claimed_items:
                        claimed_test = item.get("Claimed test", "")
                        date = item.get("Date")
                        # Try to find a matching CP method (fuzzy: check if any code word from claim is in CP methods)
                        code_part = claimed_test.split(" - ")[0].strip() if " - " in claimed_test else claimed_test
                        matched = any(code_part in m for m in cp_methods)
                        discrepancy_rows.append({
                            "TR": tr,
                            "Date": date,
                            "Match": "Yes" if matched else "No",
                            "Claimed test": claimed_test,
                            "CivilPro test": "; ".join(cp_methods) if not matched else "Matched",
                        })

        disc_df = pd.DataFrame(discrepancy_rows) if discrepancy_rows else pd.DataFrame(
            columns=["TR", "Date", "Match", "Claimed test", "CivilPro test"]
        )

        # Write Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # All tests
            cp_df.to_excel(writer, sheet_name="All tests", index=False)

            # Tests summary (filtered)
            pivot_df.to_excel(writer, sheet_name="Tests summary", index=False)
            ws_summ = writer.sheets["Tests summary"]
            ws_summ.cell(row=1, column=3).value = f"Filter: {filter_start} to {filter_end}"

            # Discrepancies
            disc_df.to_excel(writer, sheet_name="Discrepancies", index=False)

            # Auto-fit columns
            for sheet_name in writer.sheets:
                ws_sheet = writer.sheets[sheet_name]
                for col_cells in ws_sheet.columns:
                    max_length = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            max_length = max(max_length, len(str(cell.value or "")))
                        except Exception:
                            pass
                    ws_sheet.column_dimensions[col_letter].width = min(max_length + 2, 60)

        output.seek(0)
        st.download_button(
            label="Download Excel Report",
            data=output,
            file_name="tr_claim_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Preview discrepancies
        if not disc_df.empty:
            st.subheader("Discrepancy Preview")
            no_match = disc_df[disc_df["Match"] != "Yes"]
            st.write(f"{len(no_match)} line item(s) with no match or issues:")
            st.dataframe(no_match)
        elif claim_df is not None:
            st.success("No discrepancies found — all claimed TRs matched CivilPro data.")
        else:
            st.info("No claim sheet uploaded — discrepancies tab will be empty.")
else:
    st.info("Process your CivilPro PDFs in Step 1 first, then come back here to download.")
